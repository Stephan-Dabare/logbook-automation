import datetime
import requests
import json
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.drawing.image import Image
import os
import io
from typing import List, Dict, Tuple

# --- OLLAMA LLM CONFIGURATION ---
OLLAMA_MODEL = "gemma3:4b"
OLLAMA_HOST = "http://localhost:11434"

logger = logging.getLogger("logbook.generator")

def get_activity_num_with_ollama(task_description, activity_list, model=OLLAMA_MODEL, host=OLLAMA_HOST):
    """
    Uses the LLM to find matching activity numbers for a given task description.
    """
    if not activity_list:
        return "N/A"

    activities_str = "\n".join(activity_list)
    logger.info("Requesting activity numbers (task len=%d)", len(task_description))
    prompt = f"""
    You are a precise project management assistant. Your task is to analyze a work description and identify the MOST RELEVANT activities from the provided list.
    
    IMPORTANT RULES:
    1. You MUST select between 2 and 6 activity numbers (minimum 2, maximum 6).
    2. Choose only the activities that are directly relevant to the work described.
    3. Rank them by relevance and select the top 2-6 matches.
    4. Respond with ONLY the activity numbers, separated by a comma and a space.
    5. Do not add any explanation or other text.

    Here is the list of official activities:
    ---
    {activities_str}
    ---

    Now, determine between 2 and 6 most relevant activity numbers for the following work description:
    "{task_description}"
    """

    try:
        payload = {
            "model": model,
            "prompt": prompt,
            "stream": False
        }
        response = requests.post(f"{host}/api/generate", json=payload, timeout=300)
        response.raise_for_status()
        response_data = response.json()
        activity_num = response_data.get('response', 'N/A').strip()
        logger.info("LLM activity response: %s", activity_num[:80])

        if activity_num != 'N/A':
            activities = [a.strip() for a in activity_num.split(',') if a.strip()]
            if len(activities) > 6:
                activities = activities[:6]
            activity_num = ", ".join(activities)

        return activity_num

    except requests.exceptions.RequestException as e:
        logger.error("LLM error while fetching activity numbers: %s", e)
        print(f"  - LLM Error (Activity No.): {e}")
        return "N/A"

def generate_summary_with_ollama(tasks_for_week, model=OLLAMA_MODEL, host=OLLAMA_HOST):
    """
    Sends weekly tasks to a local LLM to generate a summary of problems and solutions.
    """
    if not tasks_for_week.strip():
        return "No specific problems noted.", "Solutions were implemented as part of the tasks."

    logger.info("Generating summary for week tasks (%d chars)", len(tasks_for_week))
    prompt = f"""
    Based on the following list of tasks completed in a week, act as a project manager's assistant 
    to infer one potential problem or challenge and one corresponding solution.
    Your response MUST be a single, valid JSON object with two keys: "problems_encountered" and "solutions_found".
    Do not add any text before or after the JSON object.

    Tasks for the week:
    ---
    {tasks_for_week}
    ---
    """
    try:
        payload = {"model": model, "prompt": prompt, "format": "json", "stream": False}
        response = requests.post(f"{host}/api/generate", json=payload, timeout=300)
        response.raise_for_status()
        response_data = response.json()
        llm_output = json.loads(response_data.get('response', '{}'))
        problems = llm_output.get("problems_encountered", "Could not generate summary.")
        solutions = llm_output.get("solutions_found", "Could not generate summary.")
        return problems, solutions
    except Exception as e:
        logger.error("Summary generation failed: %s", e)
        print(f"Error generating summary: {e}")
        return "Error generating summary.", "Please check LLM connection."

def generate_supervisor_comment_with_ollama(tasks_for_week, model=OLLAMA_MODEL, host=OLLAMA_HOST):
    """
    Generates a professional supervisor comment based on the week's tasks.
    """
    if not tasks_for_week.strip():
        return "No tasks recorded for this week."

    logger.info("Generating supervisor comment (%d chars)", len(tasks_for_week))
    prompt = f"""
    You are an Industrial Supervisor reviewing a student's weekly log book.
    Based on the following tasks completed by the student this week, write a brief, professional, and encouraging comment approving their work.
    The comment should be 1-2 sentences long.
    
    Tasks:
    {tasks_for_week}
    """
    try:
        payload = {"model": model, "prompt": prompt, "stream": False}
        response = requests.post(f"{host}/api/generate", json=payload, timeout=300)
        response.raise_for_status()
        comment = response.json().get('response', 'Good progress this week.').strip()
        logger.info("Supervisor comment generated")
        return comment
    except Exception as e:
        logger.error("Supervisor comment generation failed: %s", e)
        print(f"Error generating supervisor comment: {e}")
        return "Good progress this week."

def read_tasks_from_sheet(wb, task_sheet_name):
    tasks = {}
    if task_sheet_name in wb.sheetnames:
        ws_tasks = wb[task_sheet_name]
        for row in ws_tasks.iter_rows(min_row=2, values_only=True):
            date_val, task_description = row[0], row[1]
            if date_val and task_description:
                task_date = date_val.date() if isinstance(date_val, datetime.datetime) else None
                if task_date:
                    tasks[task_date] = str(task_description).strip()
    return tasks

def read_activity_nums_from_sheet(wb, activity_nums_sheet_name):
    activities = []
    if activity_nums_sheet_name in wb.sheetnames:
        ws_activities = wb[activity_nums_sheet_name]
        for row in ws_activities.iter_rows(min_row=1, values_only=True):
            if len(row) >= 3:
                num, desc = row[1], row[2]
                if num and desc and '.' in str(num):
                    activities.append(f"{num} {desc}")
    return activities

def parse_excel_to_weeks(file_content: bytes, start_date_str: str, end_date_str: str) -> List[Dict]:
    """
    Parses the uploaded Excel file and returns a list of weekly data structures.
    """
    try:
        logger.info("Parsing Excel between %s and %s", start_date_str, end_date_str)
        wb = load_workbook(filename=io.BytesIO(file_content))
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        
        tasks_data = read_tasks_from_sheet(wb, "task_sheet")
        activity_nums_data = read_activity_nums_from_sheet(wb, "activity_nums")
        
        weeks = []
        current_date = start_date
        
        # Align to the start of the week (Monday) if needed, or just iterate
        # Logic: Iterate day by day. When Sunday is reached, close the week.
        
        current_week_tasks = []
        current_week_summary_text = ""
        
        while current_date <= end_date:
            day_index = current_date.weekday() # 0=Mon, 6=Sun
            
            if current_date in tasks_data:
                task_desc = tasks_data[current_date]
                activity_num = get_activity_num_with_ollama(task_desc, activity_nums_data)
                current_week_tasks.append({
                    "date": current_date.strftime("%Y-%m-%d"),
                    "description": task_desc,
                    "activity_no": activity_num
                })
                current_week_summary_text += f"- {task_desc}\n"
            
            if day_index == 6 or current_date == end_date:
                # End of week
                if current_week_tasks:
                    problems, solutions = generate_summary_with_ollama(current_week_summary_text)
                    weeks.append({
                        "week_ending": current_date.strftime("%Y-%m-%d"),
                        "tasks": current_week_tasks,
                        "tasks_summary_text": current_week_summary_text,
                        "problems": problems,
                        "solutions": solutions,
                        "supervisor_comment": "" # To be filled later
                    })
                current_week_tasks = []
                current_week_summary_text = ""
            
            current_date += datetime.timedelta(days=1)
            
        logger.info("Parsed %d weeks", len(weeks))
        return weeks
    except Exception as e:
        logger.exception("Failed to parse Excel")
        print(f"Error parsing Excel: {e}")
        raise e

def create_final_excel(weeks_data: List[Dict], signature_img_bytes: bytes = None) -> io.BytesIO:
    """
    Generates the final Excel file from the approved weekly data.
    """
    logger.info("Building final workbook for %d weeks (signature=%s)", len(weeks_data), bool(signature_img_bytes))
    wb = Workbook()
    ws = wb.active
    ws.title = "log"
    
    # Column dimensions
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 20
    
    row_offset = 1
    
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for week in weeks_data:
        # Header
        ws[f'A{row_offset}'] = "WEEK ENDING"
        ws[f'A{row_offset}'].font = bold_font
        ws[f'B{row_offset}'] = week['week_ending']
        ws[f'B{row_offset}'].font = bold_font
        
        headers = {'A': "DAYS", 'B': "DATE", 'C': "DESCRIPTION OF WORK CARRIED OUT", 'D': "ACTIVITY NO."}
        header_row = row_offset + 1
        for col, text in headers.items():
            cell = ws[f'{col}{header_row}']
            cell.value = text
            cell.font = bold_font
            cell.alignment = center_align
            cell.border = thin_border
            
        days_of_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
        week_start_date = datetime.datetime.strptime(week['week_ending'], "%Y-%m-%d").date() - datetime.timedelta(days=6)
        
        # Fill days
        task_map = {t['date']: t for t in week['tasks']}
        
        for i, day in enumerate(days_of_week):
            current_row = row_offset + 2 + i
            date_obj = week_start_date + datetime.timedelta(days=i)
            date_str = date_obj.strftime("%Y-%m-%d")
            
            ws[f'A{current_row}'].value = day
            ws[f'A{current_row}'].border = thin_border
            ws[f'B{current_row}'].border = thin_border
            ws[f'C{current_row}'].border = thin_border
            ws[f'D{current_row}'].border = thin_border
            
            if date_str in task_map:
                task = task_map[date_str]
                ws[f'B{current_row}'].value = date_str
                ws[f'C{current_row}'].value = task['description']
                ws[f'C{current_row}'].alignment = Alignment(wrap_text=True)
                ws[f'D{current_row}'].value = task['activity_no']
                ws[f'D{current_row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Problems/Solutions
        ws[f'C{row_offset + 9}'].value = "PROBLEMS ENCOUNTERED"
        ws[f'C{row_offset + 9}'].font = bold_font
        ws[f'C{row_offset + 9}'].border = thin_border
        
        ws[f'D{row_offset + 9}'].value = "SOLUTIONS FOUND"
        ws[f'D{row_offset + 9}'].font = bold_font
        ws[f'D{row_offset + 9}'].border = thin_border
        
        ws[f'C{row_offset + 10}'].value = week['problems']
        ws[f'C{row_offset + 10}'].alignment = Alignment(wrap_text=True)
        ws[f'C{row_offset + 10}'].border = thin_border
        
        ws[f'D{row_offset + 10}'].value = week['solutions']
        ws[f'D{row_offset + 10}'].alignment = Alignment(wrap_text=True)
        ws[f'D{row_offset + 10}'].border = thin_border
        
        # Supervisor Comments
        ws.merge_cells(f'A{row_offset + 11}:D{row_offset + 11}')
        ws[f'A{row_offset + 11}'].value = "INDUSTRIAL SUPERVISOR'S COMMENTS"
        ws[f'A{row_offset + 11}'].font = bold_font
        ws[f'A{row_offset + 11}'].border = thin_border
        
        ws.merge_cells(f'A{row_offset + 12}:D{row_offset + 12}')
        ws[f'A{row_offset + 12}'].value = week.get('supervisor_comment', '')
        ws[f'A{row_offset + 12}'].alignment = Alignment(wrap_text=True)
        ws[f'A{row_offset + 12}'].border = thin_border
        
        # Designation & Signature
        ws.merge_cells(f'A{row_offset + 13}:B{row_offset + 13}')
        ws[f'A{row_offset + 13}'].value = "DESIGNATION\nIndustrial Supervisor"
        ws[f'A{row_offset + 13}'].font = bold_font
        ws[f'A{row_offset + 13}'].alignment = center_align
        ws[f'A{row_offset + 13}'].border = thin_border
        ws[f'B{row_offset + 13}'].border = thin_border # Merged
        
        ws.merge_cells(f'C{row_offset + 13}:D{row_offset + 13}')
        ws[f'C{row_offset + 13}'].value = "SIGNATURE" # Placeholder if no image
        ws[f'C{row_offset + 13}'].font = bold_font
        ws[f'C{row_offset + 13}'].alignment = center_align
        ws[f'C{row_offset + 13}'].border = thin_border
        ws[f'D{row_offset + 13}'].border = thin_border # Merged
        
        ws.row_dimensions[row_offset + 13].height = 40
        
        if signature_img_bytes:
            try:
                img = Image(io.BytesIO(signature_img_bytes))
                img.width = 120
                img.height = 35
                img.anchor = f'C{row_offset + 13}'
                ws.add_image(img)
                ws[f'C{row_offset + 13}'].value = "" # Clear text
            except Exception as e:
                logger.error("Failed to add signature image: %s", e)
                print(f"Error adding signature: {e}")
        
        row_offset += 16
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    logger.info("Workbook ready (%d bytes)", output.getbuffer().nbytes)
    return output
