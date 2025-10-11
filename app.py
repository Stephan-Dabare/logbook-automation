import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- CONFIGURATION ---
# This is the Excel file that will always be edited.
EXCEL_FILE_PATH = "my_record_book.xlsx"


def create_table_structure(ws, start_row):
    """
    Creates the static structure, headers, and styling for a single weekly table.

    Args:
        ws (Worksheet): The openpyxl worksheet object.
        start_row (int): The row number to start creating the table from.
    """

    # --- Styles ---
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')

    # Define border styles
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # --- Header Rows ---
    headers = {
        'A': "DAYS",
        'B': "DATE",
        'C': "DESCRIPTION OF WORK CARRIED OUT",
        'D': "ACTIVITY NO."
    }

    ws[f'A{start_row}'] = "WEEK ENDING"
    ws[f'A{start_row}'].font = bold_font

    header_row = start_row + 1
    for col_letter, header_text in headers.items():
        cell = ws[f'{col_letter}{header_row}']
        cell.value = header_text
        cell.font = bold_font
        cell.alignment = center_align

    # --- Static Day Labels ---
    days_of_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
    for i, day in enumerate(days_of_week):
        cell = ws[f'A{start_row + 2 + i}']
        cell.value = day
        cell.alignment = left_align

    # --- Footer Section ---
    problems_cell = ws[f'C{start_row + 9}']
    problems_cell.value = "PROBLEMS ENCOUNTERED"
    problems_cell.font = bold_font
    problems_cell.alignment = center_align

    solutions_cell = ws[f'D{start_row + 9}']
    solutions_cell.value = "SOLUTIONS FOUND"
    solutions_cell.font = bold_font
    solutions_cell.alignment = center_align

    ws.merge_cells(f'A{start_row + 11}:D{start_row + 11}')
    supervisor_cell = ws[f'A{start_row + 11}']
    supervisor_cell.value = "INDUSTRIAL SUPERVISOR'S COMMENTS"
    supervisor_cell.font = bold_font
    supervisor_cell.alignment = center_align

    ws.merge_cells(f'A{start_row + 12}:D{start_row + 12}')

    ws.merge_cells(f'A{start_row + 13}:B{start_row + 13}')
    designation_cell = ws[f'A{start_row + 13}']
    designation_cell.value = "DESIGNATION"
    designation_cell.font = bold_font
    designation_cell.alignment = center_align

    ws.merge_cells(f'C{start_row + 13}:D{start_row + 13}')

    # --- Apply Borders to the entire table structure ---
    for row in ws.iter_rows(min_row=start_row, max_row=start_row + 13, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border


def read_tasks_from_sheet(wb, task_sheet_name):
    """
    Reads tasks from a specified sheet and maps them to dates.

    Args:
        wb (Workbook): The openpyxl workbook object.
        task_sheet_name (str): The name of the sheet containing tasks.

    Returns:
        dict: A dictionary mapping dates to task descriptions.
    """
    tasks = {}
    if task_sheet_name in wb.sheetnames:
        ws_tasks = wb[task_sheet_name]
        for row in ws_tasks.iter_rows(min_row=2, values_only=True):
            date_val, task_description = row[0], row[1]
            if date_val and task_description:
                task_date = date_val.date() if isinstance(date_val, datetime.datetime) else None
                if task_date:
                    tasks[task_date] = str(task_description).strip()
    else:
        print(f"Warning: Task sheet '{task_sheet_name}' not found. No tasks will be populated.")
    return tasks


def generate_weekly_report(start_date_str, end_date_str, sheet_name, task_sheet_name):
    """
    Generates or updates an Excel report with weekly tables and populates tasks.
    """
    tasks_data = {}
    try:
        start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        print("Error: Please use the YYYY-MM-DD format for dates.")
        return

    try:
        wb = load_workbook(EXCEL_FILE_PATH)
        tasks_data = read_tasks_from_sheet(wb, task_sheet_name)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 20
        print(f"Info: '{EXCEL_FILE_PATH}' not found. A new file will be created.")

    row_offset = ws.max_row + 3 if ws.max_row > 1 else 1

    current_date = start_date
    is_new_week = True

    while current_date <= end_date:
        if is_new_week:
            create_table_structure(ws, row_offset)
            is_new_week = False

        day_index = current_date.weekday()
        current_data_row = row_offset + 2 + day_index

        date_cell = ws[f'B{current_data_row}']
        date_cell.value = current_date.strftime("%Y-%m-%d")
        date_cell.alignment = Alignment(horizontal='center', vertical='center')

        # --- Populate task description from the tasks data ---
        if current_date in tasks_data:
            task_cell = ws[f'C{current_data_row}']
            task_cell.value = tasks_data[current_date]
            task_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        if day_index == 6:
            week_ending_cell = ws[f'B{row_offset}']
            week_ending_cell.value = current_date.strftime("%Y-%m-%d")
            week_ending_cell.font = Font(bold=True)
            week_ending_cell.alignment = Alignment(horizontal='left', vertical='center')

            row_offset += 16
            is_new_week = True

        current_date += datetime.timedelta(days=1)

    last_processed_day = current_date - datetime.timedelta(days=1)
    if last_processed_day.weekday() != 6:
        week_ending_date = last_processed_day + datetime.timedelta(days=(6 - last_processed_day.weekday()))
        week_ending_cell = ws[f'B{row_offset}']
        week_ending_cell.value = week_ending_date.strftime("%Y-%m-%d")
        week_ending_cell.font = Font(bold=True)
        week_ending_cell.alignment = Alignment(horizontal='left', vertical='center')

    wb.save(EXCEL_FILE_PATH)
    print(f"Successfully updated report: {EXCEL_FILE_PATH} (Sheet: {sheet_name})")


if __name__ == "__main__":
    # --- User Configuration ---
    START_DATE = "2025-05-15"
    END_DATE = "2025-09-30"
    SHEET_NAME = "log2"
    TASK_SHEET_NAME = "task_sheet"

    generate_weekly_report(START_DATE, END_DATE, SHEET_NAME, TASK_SHEET_NAME)

